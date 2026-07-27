from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.interfaces.reporter import ReportGeneratorInterface
from app.dto.report import EvaluationReportDTO
from app.utils.logger import get_logger

logger = get_logger("excel_reporter")

class ExcelReportGenerator(ReportGeneratorInterface):
    """
    Generates enterprise Excel evaluation reports with openpyxl:
    - Summary Dashboard tab
    - Test Case Results tab
    - Metric Breakdown tab
    - Professional styling, custom header fills, fonts, and conditional formatting.
    """

    async def generate_report(self, report: EvaluationReportDTO, output_dir: str) -> str:
        out_path = Path(output_dir)
        out_path.mkdir(parents=True, exist_ok=True)
        file_path = out_path / "evaluation_report.xlsx"

        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        # Color Palette & Styles
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        sub_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        pass_font = Font(color="006100", bold=True)
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        fail_font = Font(color="9C0006", bold=True)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        # -------------------------------------------------------------
        # SHEET 1: Summary Dashboard
        # -------------------------------------------------------------
        ws_summary = wb.create_sheet(title="Executive Summary")
        ws_summary.views.sheetView[0].showGridLines = True

        ws_summary.append(["Agentic AI Evaluation Executive Summary"])
        ws_summary.merge_cells("A1:D1")
        ws_summary["A1"].font = Font(name="Calibri", size=16, bold=True, color="1F4E79")

        ws_summary.append([])
        ws_summary.append(["Metric Key", "Metric Value"])
        for cell in ws_summary[3]:
            cell.fill = header_fill
            cell.font = header_font

        summary_rows = [
            ("Target System Under Test", report.target_agent),
            ("Report Timestamp", report.timestamp),
            ("Total Test Cases Evaluated", report.total_test_cases),
            ("Passed Test Cases", report.passed_test_cases),
            ("Failed Test Cases", report.failed_test_cases),
            ("Pass Rate Percentage", f"{report.pass_rate_percentage:.1f}%"),
            ("Overall Weighted Score (0-100)", f"{report.overall_weighted_score:.2f}"),
            ("Total Rule Violations", report.total_violations_count),
            ("Total Duration (ms)", f"{report.total_duration_ms:.2f} ms"),
        ]

        for key, val in summary_rows:
            ws_summary.append([key, val])

        ws_summary.append([])
        ws_summary.append(["Evaluation Dimension", "Average Score (0-100)", "Dimension Weight", "Weighted Contribution"])
        for cell in ws_summary[15]:
            cell.fill = sub_header_fill
            cell.font = Font(name="Calibri", size=11, bold=True, color="1F4E79")

        for dim in report.dimension_summaries:
            ws_summary.append([
                dim.evaluator_name.capitalize(),
                round(dim.average_score, 2),
                f"{dim.weight * 100:.0f}%",
                round(dim.weighted_average_score, 2)
            ])

        # -------------------------------------------------------------
        # SHEET 2: Detailed Test Cases
        # -------------------------------------------------------------
        ws_details = wb.create_sheet(title="Test Case Results")
        ws_details.views.sheetView[0].showGridLines = True

        headers = ["Test Case ID", "Status", "Weighted Score", "Raw Score", "Execution Time (ms)", "Violations Count", "Violations Details"]
        ws_details.append(headers)
        for cell in ws_details[1]:
            cell.fill = header_fill
            cell.font = header_font

        for res in report.results:
            status_str = "PASSED" if res.passed else "FAILED"
            violations_str = " | ".join(res.violations) if res.violations else "None"
            row = [
                res.test_case_id,
                status_str,
                res.weighted_score,
                res.overall_score,
                res.execution_time_ms,
                len(res.violations),
                violations_str
            ]
            ws_details.append(row)
            current_row = ws_details.max_row
            status_cell = ws_details.cell(row=current_row, column=2)
            if res.passed:
                status_cell.fill = pass_fill
                status_cell.font = pass_font
            else:
                status_cell.fill = fail_fill
                status_cell.font = fail_font

        # -------------------------------------------------------------
        # SHEET 3: Dimension Breakdown
        # -------------------------------------------------------------
        ws_metrics = wb.create_sheet(title="Dimension Metrics")
        ws_metrics.views.sheetView[0].showGridLines = True

        metric_headers = ["Test Case ID", "Evaluator Dimension", "Score (0-100)", "Weight", "Weighted Score", "Violations"]
        ws_metrics.append(metric_headers)
        for cell in ws_metrics[1]:
            cell.fill = header_fill
            cell.font = header_font

        for res in report.results:
            for dim_name, m_score in res.metric_scores.items():
                v_str = " | ".join(m_score.violations) if m_score.violations else "None"
                ws_metrics.append([
                    res.test_case_id,
                    dim_name.capitalize(),
                    m_score.score,
                    m_score.weight,
                    m_score.weighted_score,
                    v_str
                ])

        # Auto-adjust column widths for all sheets
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                    cell.border = thin_border
                sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)

        wb.save(file_path)
        logger.info(f"Excel evaluation report saved to: {file_path.resolve()}")
        return str(file_path.resolve())
